"""Export service — generates Excel, CSV, PDF files from report data."""

from __future__ import annotations

import csv
import io
import uuid
from pathlib import Path
from xml.sax.saxutils import escape

from sqlalchemy.ext.asyncio import AsyncSession

from app.services.report_builder import build_report


async def export_report(
    db: AsyncSession,
    project_id: uuid.UUID,
    report_type: str,
    fmt: str,
) -> io.BytesIO | None:
    # For "issues" report type, build it separately
    if report_type == "issues":
        data = await _build_issues_report(db, project_id)
    else:
        data = await build_report(db, project_id, report_type)

    if data is None:
        return None

    if fmt == "pdf" and "sections" in data:
        return _sections_to_pdf(data.get("title", report_type), data["sections"])

    # Summary report has "sections" instead of flat columns/rows
    if "sections" in data:
        columns = []
        rows = []
        for section in data["sections"]:
            if rows:
                rows.append([])  # separator
            rows.append([section["title"]])
            rows.append(section["columns"])
            rows.extend(section["rows"])
        title = data.get("title", report_type)
    else:
        columns = data["columns"]
        rows = data["rows"]
        title = data.get("title", report_type)

    if fmt == "xlsx":
        return _to_xlsx(title, columns, rows)
    elif fmt == "csv":
        return _to_csv(columns, rows)
    elif fmt == "pdf":
        return _to_pdf(title, columns, rows)
    return None


def _to_xlsx(title: str, columns: list, rows: list) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    # Title row
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(columns), 1))
    ws.cell(1, 1).font = Font(bold=True, size=14)
    ws.append([])

    # Header row
    if columns:
        ws.append(columns)
        header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(3, col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row in rows:
        ws.append([str(v) if v is not None else "" for v in row])

    # Auto-width
    for col in ws.columns:
        max_len = 0
        column_letter = None
        for cell in col:
            if hasattr(cell, 'column_letter'):
                column_letter = cell.column_letter
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        if column_letter:
            ws.column_dimensions[column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _to_csv(columns: list, rows: list) -> io.BytesIO:
    buf = io.StringIO()
    writer = csv.writer(buf)
    if columns:
        writer.writerow(columns)
    for row in rows:
        writer.writerow(row)
    result = io.BytesIO(buf.getvalue().encode("utf-8-sig"))
    result.seek(0)
    return result


def _to_pdf(title: str, columns: list, rows: list) -> io.BytesIO:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import LongTable, Paragraph, SimpleDocTemplate, Spacer, TableStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    font_name, font_bold = _register_pdf_fonts(pdfmetrics, TTFont)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "PdfTitle",
        parent=styles["Title"],
        fontName=font_bold,
        fontSize=16,
        leading=20,
        spaceAfter=6 * mm,
    )
    header_style = ParagraphStyle(
        "PdfHeader",
        parent=styles["Normal"],
        fontName=font_bold,
        fontSize=9,
        leading=11,
        textColor=colors.white,
        alignment=1,
        wordWrap="CJK",
    )
    body_style = ParagraphStyle(
        "PdfBody",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=8,
        leading=10,
        wordWrap="CJK",
    )

    story = []

    story.append(Paragraph(title, title_style))

    if not columns and not rows:
        story.append(Paragraph("Нет данных", body_style))
        doc.build(story)
        buf.seek(0)
        return buf

    table = _build_pdf_table(
        columns=columns,
        rows=rows,
        available_width=doc.width,
        header_style=header_style,
        body_style=body_style,
        colors_module=colors,
        long_table_cls=LongTable,
        table_style_cls=TableStyle,
    )
    if table is None:
        story.append(Paragraph("Нет данных", body_style))
        doc.build(story)
        buf.seek(0)
        return buf

    story.append(table)

    doc.build(story)
    buf.seek(0)
    return buf


def _sections_to_pdf(title: str, sections: list[dict]) -> io.BytesIO:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import LongTable, Paragraph, SimpleDocTemplate, Spacer, TableStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    font_name, font_bold = _register_pdf_fonts(pdfmetrics, TTFont)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "PdfTitle",
        parent=styles["Title"],
        fontName=font_bold,
        fontSize=16,
        leading=20,
        spaceAfter=6 * mm,
    )
    section_style = ParagraphStyle(
        "PdfSection",
        parent=styles["Heading2"],
        fontName=font_bold,
        fontSize=12,
        leading=14,
        spaceBefore=2 * mm,
        spaceAfter=3 * mm,
    )
    header_style = ParagraphStyle(
        "PdfHeader",
        parent=styles["Normal"],
        fontName=font_bold,
        fontSize=9,
        leading=11,
        textColor=colors.white,
        alignment=1,
        wordWrap="CJK",
    )
    body_style = ParagraphStyle(
        "PdfBody",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=8,
        leading=10,
        wordWrap="CJK",
    )

    story = [Paragraph(title, title_style)]

    if not sections:
        story.append(Paragraph("Нет данных", body_style))
        doc.build(story)
        buf.seek(0)
        return buf

    for idx, section in enumerate(sections):
        story.append(Paragraph(section.get("title", "Раздел"), section_style))
        table = _build_pdf_table(
            columns=section.get("columns", []),
            rows=section.get("rows", []),
            available_width=doc.width,
            header_style=header_style,
            body_style=body_style,
            colors_module=colors,
            long_table_cls=LongTable,
            table_style_cls=TableStyle,
        )
        if table is None:
            story.append(Paragraph("Нет данных", body_style))
        else:
            story.append(table)
        if idx < len(sections) - 1:
            story.append(Spacer(1, 5 * mm))

    doc.build(story)
    buf.seek(0)
    return buf


def _register_pdf_fonts(pdfmetrics, tt_font_cls) -> tuple[str, str]:
    import reportlab

    font_dir = Path(__file__).resolve().parent.parent / "fonts"
    reportlab_font_dir = Path(reportlab.__file__).resolve().parent / "fonts"

    candidates = [
        (Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"), Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"), "SysDejaVuSans", "SysDejaVuSans-Bold"),
        (Path("/System/Library/Fonts/Supplemental/Arial Unicode.ttf"), Path("/System/Library/Fonts/Supplemental/Arial Bold.ttf"), "ArialUnicodeMS", "ArialUnicodeMS-Bold"),
        (Path("/System/Library/Fonts/Supplemental/Arial.ttf"), Path("/System/Library/Fonts/Supplemental/Arial Bold.ttf"), "Arial", "Arial-Bold"),
        (font_dir / "DejaVuSans.ttf", font_dir / "DejaVuSans-Bold.ttf", "DejaVuSans", "DejaVuSans-Bold"),
        (reportlab_font_dir / "Vera.ttf", reportlab_font_dir / "VeraBd.ttf", "Vera", "VeraBd"),
    ]

    registered = set(pdfmetrics.getRegisteredFontNames())
    for regular_path, bold_path, regular_name, bold_name in candidates:
        if not regular_path.exists() or not bold_path.exists():
            continue
        try:
            if regular_name not in registered:
                pdfmetrics.registerFont(tt_font_cls(regular_name, str(regular_path)))
            if bold_name not in registered:
                pdfmetrics.registerFont(tt_font_cls(bold_name, str(bold_path)))
            return regular_name, bold_name
        except Exception:
            continue

    return "Helvetica", "Helvetica-Bold"


def _build_pdf_table(
    columns: list,
    rows: list,
    available_width: float,
    header_style,
    body_style,
    colors_module,
    long_table_cls,
    table_style_cls,
):
    from reportlab.platypus import Paragraph

    normalized_rows = [["" if value is None else str(value) for value in row] for row in rows if row]
    if not columns and not normalized_rows:
        return None

    col_count = max(len(columns), max((len(row) for row in normalized_rows), default=0), 1)
    normalized_columns = [str(value) for value in columns] + [""] * (col_count - len(columns))
    padded_rows = [row + [""] * (col_count - len(row)) for row in normalized_rows]

    col_widths = _compute_pdf_col_widths(normalized_columns, padded_rows, available_width)
    table_data = []

    if columns:
        table_data.append([Paragraph(escape(value), header_style) for value in normalized_columns])

    for row in padded_rows:
        table_data.append([Paragraph(escape(value).replace("\n", "<br/>"), body_style) for value in row])

    if not table_data:
        return None

    table = long_table_cls(table_data, colWidths=col_widths, repeatRows=1 if columns else 0)
    style_commands = [
        ("GRID", (0, 0), (-1, -1), 0.5, colors_module.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
    ]
    if columns:
        style_commands.extend([
            ("BACKGROUND", (0, 0), (-1, 0), colors_module.HexColor("#2E75B6")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors_module.white),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors_module.white, colors_module.HexColor("#F0F4F8")]),
        ])
    else:
        style_commands.append(("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors_module.white, colors_module.HexColor("#F0F4F8")]))
    table.setStyle(table_style_cls(style_commands))
    return table


def _compute_pdf_col_widths(columns: list[str], rows: list[list[str]], available_width: float) -> list[float]:
    col_count = max(len(columns), max((len(row) for row in rows), default=0), 1)
    if col_count == 1:
        return [available_width]

    samples = [columns, *rows[:50]]
    weights = []
    for idx in range(col_count):
        values = [row[idx] for row in samples if idx < len(row)]
        max_len = max((len(value) for value in values), default=8)
        avg_len = sum(len(value) for value in values) / max(len(values), 1)
        weights.append(max(1.0, min(max_len, 40)) + min(avg_len, 20))

    total_weight = sum(weights) or col_count
    min_width = min(available_width / col_count * 0.65, 55)
    widths = [max(min_width, available_width * weight / total_weight) for weight in weights]
    total_width = sum(widths)

    if total_width > available_width:
        scale = available_width / total_width
        widths = [width * scale for width in widths]

    return widths


async def _build_issues_report(db: AsyncSession, project_id: uuid.UUID) -> dict | None:
    from sqlalchemy import select
    from app.models.project import Issue, Project

    project = await db.get(Project, project_id)
    if not project:
        return None

    stmt = (
        select(Issue)
        .where(Issue.project_id == project_id)
        .order_by(Issue.severity, Issue.created_at.desc())
    )
    result = await db.execute(stmt)
    issues = result.scalars().all()

    rows = []
    for i, issue in enumerate(issues, 1):
        rows.append([
            i,
            issue.severity,
            issue.category,
            issue.message,
            issue.status,
            issue.created_at.strftime("%Y-%m-%d %H:%M") if issue.created_at else "—",
        ])

    return {
        "title": "Журнал замечаний",
        "columns": ["№", "Критичность", "Категория", "Описание", "Статус", "Дата"],
        "rows": rows,
    }
